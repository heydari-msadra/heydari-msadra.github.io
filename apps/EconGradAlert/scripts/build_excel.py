#!/usr/bin/env python3
"""Regenerate data/tracker.xlsx from data/tracker.json.

tracker.json is the single source of truth; this script is the one place
that turns it into the downloadable workbook. Run after any edit to
tracker.json — the update agent runs it as part of every scheduled pass.

Usage:
    python scripts/build_excel.py
"""
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
JSON_PATH = os.path.join(ROOT, "data", "tracker.json")
XLSX_PATH = os.path.join(ROOT, "data", "tracker.xlsx")

PROGRAM_COLUMNS = [
    ("id", "Program ID"),
    ("degree", "Degree Level"),
    ("name", "Program Name"),
    ("uni", "University"),
    ("repec_rank", "RePEc IDEAS Rank"),
    ("region", "Region"),
    ("country", "Country"),
    ("city", "City"),
    ("link", "Program Link"),
    ("ielts", "IELTS Requirement"),
    ("toefl", "TOEFL Requirement"),
    ("gre", "GRE Requirement"),
    ("other", "Other Requirements"),
    ("start", "Program Start Date"),
    ("deadline", "Application Deadline"),
    ("urgency", "Urgency"),
    ("refs", "Reference Letters Required"),
    ("appfee", "Application Fee"),
    ("progfee", "Program/Tuition Fee"),
    ("funding_status", "Scholarship/Funding Status"),
    ("funding_detail", "Funding Details"),
    ("date_added", "Date Added"),
    ("last_updated", "Last Updated"),
    ("last_change", "Last Change"),
    ("notes", "Notes / Source"),
]

CHANGELOG_COLUMNS = [
    ("date", "Date"),
    ("program_id", "Program ID"),
    ("program_name", "Program Name"),
    ("change_type", "Change Type"),
    ("details", "Details"),
]

HEADER_FONT = Font(bold=True)


def autosize(ws, max_width=60):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[letter].width = min(max(length + 2, 10), max_width)


def build():
    with open(JSON_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()

    readme = wb.active
    readme.title = "README"
    meta = data.get("meta", {})
    lines = [
        meta.get("title", "Tracker"),
        "",
        "Last full research pass: " + meta.get("last_full_update", ""),
        "Coverage: " + meta.get("coverage_note", ""),
        "",
        "This workbook is generated from data/tracker.json — do not hand-edit it,",
        "edits will be overwritten on the next run of scripts/build_excel.py.",
        "",
        "RePEc IDEAS source: " + meta.get("repec_source", ""),
        "RePEc snapshot: " + meta.get("repec_snapshot", ""),
        "Program count: " + str(meta.get("program_count", len(data.get("programs", [])))),
        "",
        "SHEETS",
        "- Programs: the full database.",
        "- Change Log: history of additions/edits, newest first.",
    ]
    for i, line in enumerate(lines, start=1):
        readme.cell(row=i, column=1, value=line)
    readme.column_dimensions["A"].width = 100

    progs = wb.create_sheet("Programs")
    for c, (_, label) in enumerate(PROGRAM_COLUMNS, start=1):
        cell = progs.cell(row=1, column=c, value=label)
        cell.font = HEADER_FONT
    for r, p in enumerate(data.get("programs", []), start=2):
        for c, (key, _) in enumerate(PROGRAM_COLUMNS, start=1):
            progs.cell(row=r, column=c, value=p.get(key, ""))
    progs.freeze_panes = "A2"
    autosize(progs)

    log = wb.create_sheet("Change Log")
    for c, (_, label) in enumerate(CHANGELOG_COLUMNS, start=1):
        cell = log.cell(row=1, column=c, value=label)
        cell.font = HEADER_FONT
    change_log = sorted(data.get("change_log", []), key=lambda e: e.get("date", ""), reverse=True)
    for r, e in enumerate(change_log, start=2):
        for c, (key, _) in enumerate(CHANGELOG_COLUMNS, start=1):
            log.cell(row=r, column=c, value=e.get(key, ""))
    log.freeze_panes = "A2"
    autosize(log)

    wb.save(XLSX_PATH)
    print(f"Wrote {XLSX_PATH}: {len(data.get('programs', []))} programs, {len(change_log)} change_log entries")


if __name__ == "__main__":
    build()
